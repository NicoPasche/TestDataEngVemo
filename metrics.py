import pandas as pd
from fetch_data import fetch_data
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def metrics():
    # Obtiene los datos
    countries = fetch_data()

    # Crea un DataFrame con los datos de los países
    df = pd.DataFrame(countries)

    # Inicializa un diccionario para almacenar las sumas de población por continente
    pop_by_continent = {continent: 0 for continent in df['continent'].unique()}

    # Inicializa un diccionario para almacenar las monedas por continente
    currency_by_continent = {continent: [] for continent in df['continent'].unique()}

    # Inicializa un diccionario para almacenar los idiomas por continente
    language_by_continent = {continent: [] for continent in df['continent'].unique()}

    # Inicializa un diccionario para almacenar las sumas de población por idioma y continente
    pop_by_language_and_continent = {}

    # Usa un bucle para sumar las poblaciones, recoger las monedas y los idiomas
    for i, row in df.iterrows():
        pop_by_continent[row['continent']] += row['population']
        currency_by_continent[row['continent']].append(row['currency'])
        if row['language'] is not None:  # Comprueba si el idioma no es None
            languages = row['language'].split(',')  # Divide la cadena de idiomas por la coma
            for language in languages:
                language_by_continent[row['continent']].append({'language': language, 'population': row['population']})

                # Si el continente no está en el diccionario, añade un nuevo diccionario para este continente
                if row['continent'] not in pop_by_language_and_continent:
                    pop_by_language_and_continent[row['continent']] = {}

                # Si el idioma ya está en el diccionario para este continente, suma la población. Si no, añade el idioma al diccionario con su población.
                if language in pop_by_language_and_continent[row['continent']]:
                    pop_by_language_and_continent[row['continent']][language] += row['population']
                else:
                    pop_by_language_and_continent[row['continent']][language] = row['population']

    # Calcula la moneda más común por continente
    most_common_currency_by_continent = {continent: pd.Series(currencies).mode()[0] for continent, currencies in currency_by_continent.items()}

    # Calcula el idioma más común por continente, ponderado por la población
    most_common_language_by_continent = {continent: max(languages, key=languages.get) for continent, languages in pop_by_language_and_continent.items()}

    # Crea un DataFrame con las métricas
    metrics = pd.DataFrame({
        'Población total por continente': pd.Series(pop_by_continent),
        'Número de países por continente': df['continent'].value_counts(),
        'Moneda más común por continente': pd.Series(most_common_currency_by_continent),
        'Idioma más común por continente (ponderado por población)': pd.Series(most_common_language_by_continent)
    })

    # Guarda el DataFrame en un archivo Excel
    with pd.ExcelWriter('countries.xlsx') as writer:
        df.to_excel(writer, sheet_name='Paises')
        metrics.to_excel(writer, sheet_name='Metricas')

    # Crea un gráfico de torta de la población total por continente
    plt.figure(figsize=(10, 6))
    plt.pie(metrics['Población total por continente'], labels=metrics.index, autopct='%1.1f%%')
    plt.title('Población total por continente')

    # Guarda el gráfico como una imagen
    plt.savefig('population_pie_chart.png')

    # Carga el archivo de Excel
    wb = load_workbook('countries.xlsx')

    # Selecciona la hoja 'Metricas'
    ws = wb['Metricas']

    # Carga la imagen
    img = Image('population_pie_chart.png')

    # Agrega la imagen a la hoja
    ws.add_image(img, 'A8')
    wb.save('countries.xlsx')
    # Calcula los 5 idiomas más hablados en América
    top_5_languages_in_america = sorted(pop_by_language_and_continent['Americas'].items(), key=lambda x: x[1], reverse=True)[:5]
    
    # Crea un DataFrame con los 5 idiomas más hablados en América
    df_top_5_languages_in_america = pd.DataFrame(top_5_languages_in_america, columns=['language', 'population'])
    
    # Crea un histograma con los datos
    plt.figure(figsize=(10, 6))
    plt.barh(df_top_5_languages_in_america['language'], df_top_5_languages_in_america['population'])
    plt.xlabel('Población')
    plt.ylabel('Idioma')
    plt.title('Top 5 idiomas más hablados en América')
    plt.gca().invert_yaxis()  # Invierte el eje y para que el idioma con mayor población esté en la parte superior
   # Guarda el gráfico como una imagen
    plt.savefig('top_5_languages_in_americas.png')
    
    # Carga el archivo de Excel
    wb = load_workbook('countries.xlsx')
    
    # Selecciona la hoja 'Metricas'
    ws = wb['Metricas']
    
    # Carga la imagen
    img = Image('top_5_languages_in_americas.png')
    
    # Agrega la imagen a la hoja 'Metricas'
    ws.add_image(img, 'A30')
    
    # Guarda el archivo de Excel
    wb.save('countries.xlsx')


